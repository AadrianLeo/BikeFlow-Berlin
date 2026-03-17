"""
BikeFlowBerlin V2 - Step 2: Official Berlin Bike Counter Data

Reads berlin_counters_raw.xlsx (multi-sheet structure):
  - Standortdaten : station metadata (ID, name, lat, lon)
  - Jahresdatei YYYY : hourly wide-format counts per year
    Row 1 = station_id\ndate in each column header
    Column A = hourly timestamp, all other cols = bike counts per station

Output: long-format cleaned CSV with columns:
  timestamp, station_id, station_name, bike_count, lat, lon, source
"""

import subprocess, sys

def install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for pkg in ("pandas", "openpyxl"):
    try:
        __import__(pkg)
    except ImportError:
        install(pkg)

import os
import re
import pandas as pd
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT       = os.path.dirname(SCRIPT_DIR)
RAW_DIR    = os.path.join(ROOT, "data", "official_counters", "raw")
CLEAN_DIR  = os.path.join(ROOT, "data", "official_counters", "cleaned")
os.makedirs(CLEAN_DIR, exist_ok=True)

RAW_FILE   = os.path.join(RAW_DIR, "berlin_counters_raw.xlsx")
CLEAN_FILE = os.path.join(CLEAN_DIR, "official_counters_cleaned.csv")

print("\nStep 2 - Official Berlin Bike Counter Data")
print("=" * 60)

if not os.path.exists(RAW_FILE):
    print(f"\nERROR: Raw file not found: {RAW_FILE}")
    print(f"Place berlin_counters_raw.xlsx in: {RAW_DIR}")
    sys.exit(1)

print(f"\nSource file: {os.path.basename(RAW_FILE)}")

# ── 1. Station metadata from Standortdaten ─────────────────────────────────────
# Columns: Zahlstelle | Beschreibung - Fahrtrichtung | Breitengrad | Langengrad | Installationsdatum
print("\n[1/4] Reading station metadata (Standortdaten) ...")

wb = openpyxl.load_workbook(RAW_FILE, read_only=True, data_only=True)
ws_std = wb["Standortdaten"]
standort_rows = list(ws_std.iter_rows(values_only=True))

station_info = {}
for row in standort_rows[1:]:        # skip header row
    if not row[0]:
        continue
    sid = str(row[0]).strip()
    station_info[sid] = {
        "station_name": str(row[1]).strip() if row[1] else "",
        "lat": float(row[2]) if row[2] is not None else None,
        "lon": float(row[3]) if row[3] is not None else None,
    }

print(f"   {len(station_info)} stations found")

# ── 2. Extract column headers (station IDs) from each year sheet ───────────────
# Each column header cell: "station_id\nDD.MM.YYYY"  (newline-separated)
print("\n[2/4] Reading column headers from year sheets ...")

year_sheets = [s for s in wb.sheetnames if s.startswith("Jahresdatei")]
headers_by_sheet = {}

for sheet_name in year_sheets:
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    station_ids = []
    for cell in header_row[1:]:      # skip first cell (timestamp column header)
        if cell is not None:
            # Cell format: "station_id\nDD.MM.YYYY" (newline) or "station_id DD.MM.YYYY" (space)
            # Extract just the station ID: two digits, hyphen, non-whitespace chars
            m = re.match(r"^(\d{2}-\S+)", str(cell).strip())
            sid = m.group(1) if m else str(cell).split("\n")[0].split()[0].strip()
            station_ids.append(sid)
    headers_by_sheet[sheet_name] = station_ids

wb.close()
print(f"   {len(year_sheets)} year sheets: {year_sheets[0]} to {year_sheets[-1]}")

# ── 3. Read each year sheet and melt wide -> long ──────────────────────────────
print("\n[3/4] Processing year sheets (wide -> long) ...")
all_frames = []

with pd.ExcelFile(RAW_FILE, engine="openpyxl") as xf:
    for sheet_name in year_sheets:
        station_ids = headers_by_sheet[sheet_name]

        # skiprows=1 skips the header row; header=None avoids pandas guessing headers
        df = xf.parse(sheet_name, header=None, skiprows=1)

        # Assign column names: col 0 = timestamp, rest = station IDs
        n_data_cols = df.shape[1] - 1
        col_names = ["timestamp"] + station_ids[:n_data_cols]
        # Safety pad for any extra trailing empty columns
        while len(col_names) < df.shape[1]:
            col_names.append(f"_extra_{len(col_names)}")
        df.columns = col_names

        df = df.dropna(subset=["timestamp"])

        # Melt: only known station_id columns (drop _extra_ padding)
        value_cols = [c for c in col_names[1:] if not c.startswith("_extra_")]
        df_long = df.melt(
            id_vars=["timestamp"],
            value_vars=value_cols,
            var_name="station_id",
            value_name="bike_count",
        )

        all_frames.append(df_long)
        print(f"   {sheet_name}: {len(station_ids)} stations, "
              f"{df.shape[0]:,} hours -> {len(df_long):,} rows")

# ── 4. Merge, clean, annotate, save ───────────────────────────────────────────
print("\n[4/4] Cleaning and saving ...")

df = pd.concat(all_frames, ignore_index=True)
rows_raw = len(df)
print(f"   Total rows after melt: {rows_raw:,}")

# Standardize timestamp
df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
df["timestamp"] = df["timestamp"].dt.strftime("%Y-%m-%d %H:%M:%S")

# Numeric bike_count
df["bike_count"] = pd.to_numeric(df["bike_count"], errors="coerce")

# Remove duplicates
n_before = len(df)
df = df.drop_duplicates()
print(f"   Removed {n_before - len(df):,} duplicate rows")

# Remove null / negative bike_count
n_before = len(df)
df = df[df["bike_count"].notna() & (df["bike_count"] >= 0)]
print(f"   Removed {n_before - len(df):,} null/negative rows")

# Join station name + coordinates from Standortdaten
station_df = pd.DataFrame.from_dict(station_info, orient="index").reset_index()
station_df.columns = ["station_id", "station_name", "lat", "lon"]
df = df.merge(station_df, on="station_id", how="left")

# Source tag
df["source"] = "official_berlin"

# Final column order
df = df[["timestamp", "station_id", "station_name", "bike_count", "lat", "lon", "source"]]

df.to_csv(CLEAN_FILE, index=False)

# ── Quality report ─────────────────────────────────────────────────────────────
ts_series = pd.to_datetime(df["timestamp"], errors="coerce").dropna()
n_no_coords = df["lat"].isna().sum()

print(f"\n{'=' * 60}")
print(f"Step 2 COMPLETE - Quality Report")
print(f"{'=' * 60}")
print(f"  Rows (raw after melt):    {rows_raw:,}")
print(f"  Rows (after cleaning):    {len(df):,}")
print(f"  Unique stations:          {df['station_id'].nunique()}")
print(f"  Date range:               {ts_series.min().date()} to {ts_series.max().date()}")
print(f"  Rows missing coords:      {n_no_coords:,}")
print(f"  Output: data/official_counters/cleaned/official_counters_cleaned.csv")
print(f"\n  Sample rows:")
print(df.head(8).to_string(index=False))
print(f"\nReady for Step 3 - Weather Data. Type YES to continue.")
